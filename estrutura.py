import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
import os
from PIL import Image

try:
    import openpyxl
except ImportError:
    print("Erro: A biblioteca 'openpyxl' não está instalada.")
    exit()

# ==========================================
# 1. CONFIGURAÇÕES GERAIS E SEGURANÇA
# ==========================================
ctk.set_appearance_mode("Light") 
ctk.set_default_color_theme("green") 

USUARIOS_PERMITIDOS = {
    "admin": "admin123",
    "rh.agricola": "cana2026",
    "analista": "senha123"
}

# ==========================================
# 1.5 SISTEMA DE HOVER (TOOLTIP) PARA TEXTOS LONGOS
# ==========================================
class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        if hasattr(self.widget, '_entry'):
            self.widget._entry.bind("<Enter>", self.enter)
            self.widget._entry.bind("<Leave>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.showtip)

    def unschedule(self):
        id = self.id
        self.id = None
        if id:
            self.widget.after_cancel(id)

    def showtip(self, event=None):
        text = self.widget.get()
        if not text:
            return 
            
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 2
        
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(tw, text=text, justify="left",
                      background="#1b5e20", foreground="white", relief="solid", borderwidth=1,
                      font=("Roboto", 11, "bold"))
        label.pack(ipadx=8, ipady=4)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

# ==========================================
# 2. FUNÇÃO PARA LER DO EXCEL (CASCATA)
# ==========================================
def carregar_dados_excel():
    arquivo_excel = 'parametros.xlsx'
    dados = {
        "linhas": [],          
        "requisitantes": set() 
    }
    
    if not os.path.exists(arquivo_excel):
        messagebox.showwarning("Aviso", f"Arquivo '{arquivo_excel}' não encontrado!")
        return dados

    try:
        workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
        planilha = workbook.active
        
        for linha in range(2, planilha.max_row + 1):
            val_und    = str(planilha.cell(row=linha, column=1).value or "").strip()
            val_cc     = str(planilha.cell(row=linha, column=2).value or "").strip()
            val_sub    = str(planilha.cell(row=linha, column=3).value or "").strip()
            val_gestor = str(planilha.cell(row=linha, column=4).value or "").strip()
            val_posto  = str(planilha.cell(row=linha, column=5).value or "").strip()
            val_cargo  = str(planilha.cell(row=linha, column=6).value or "").strip()
            val_req    = str(planilha.cell(row=linha, column=7).value or "").strip()
            
            if any([val_und, val_cc, val_sub, val_gestor, val_posto, val_cargo]):
                dados["linhas"].append({
                    "unidade": val_und, "cc": val_cc, "sub": val_sub,
                    "gestor": val_gestor, "posto": val_posto, "cargo": val_cargo
                })
            
            if val_req:
                dados["requisitantes"].add(val_req)
                
        dados["requisitantes"] = sorted(list(dados["requisitantes"]))
        return dados
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler Excel: {e}")
        return dados

# ==========================================
# 3. BANCO DE DADOS (SQLite - v3)
# ==========================================
def conectar_banco():
    conn = sqlite3.connect('headcount_v3.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS movimentacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_sistema TEXT, data_registro TEXT, requisitante TEXT,
            
            unidade_saida TEXT, cc_saida TEXT, subprocesso_saida TEXT,
            gestor_saida TEXT, posto_saida TEXT, cargo_saida TEXT, qtd_saida INTEGER,
            
            unidade_entrada TEXT, cc_entrada TEXT, subprocesso_entrada TEXT,
            gestor_entrada TEXT, posto_entrada TEXT, cargo_entrada TEXT, qtd_entrada INTEGER
        )
    ''')
    conn.commit()
    return conn

# ==========================================
# 4. LÓGICA DA INTERFACE (CustomTkinter)
# ==========================================
class AppHeadcount(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Movimentações - Headcount")
        self.geometry("1050x720")
        
        caminho_logo = "logo.png"
        if os.path.exists(caminho_logo):
            try:
                icone = tk.PhotoImage(file=caminho_logo)
                self.iconphoto(False, icone)
            except:
                pass

        self.usuario_logado = None
        self.qtd_saida = ctk.IntVar(value=1)
        self.qtd_entrada = ctk.IntVar(value=1)
        self.dados_excel = carregar_dados_excel()

        self.tela_login()

    def limpar_tela(self):
        for widget in self.winfo_children():
            widget.destroy()

    # --- TELA DE LOGIN ---
    def tela_login(self):
        self.limpar_tela()

        # Frame de fundo do login
        login_card = ctk.CTkFrame(self, corner_radius=25, fg_color=("#ffffff", "#2b2b2b"))
        login_card.place(relx=0.5, rely=0.5, anchor="center")

        # Inserção da logo com a margem lateral ajustada (padx=50)
        caminho_logo = "logo.png"
        if os.path.exists(caminho_logo):
            try:
                img = Image.open(caminho_logo)
                logo_image = ctk.CTkImage(light_image=img, dark_image=img, size=(150, 110))
                ctk.CTkLabel(login_card, image=logo_image, text="").pack(pady=(50, 15), padx=50) 
            except:
                ctk.CTkLabel(login_card, text="").pack(pady=(50,0), padx=50)
        else:
             ctk.CTkLabel(login_card, text="").pack(pady=(50,0), padx=50)

        # Inserção dos itens com margem lateral para empurrar o fundo branco (padx=50)
        ctk.CTkLabel(login_card, text="Movimentações\nHeadCount", font=("Roboto Bold", 28), justify="center").pack(pady=(0, 25), padx=50)
        
        self.entry_usuario = ctk.CTkEntry(login_card, placeholder_text="Usuário", width=300, height=45, font=("Roboto", 14), corner_radius=10)
        self.entry_usuario.pack(pady=10, padx=50)
        
        self.entry_senha = ctk.CTkEntry(login_card, placeholder_text="Senha", show="*", width=300, height=45, font=("Roboto", 14), corner_radius=10)
        self.entry_senha.pack(pady=10, padx=50)
        
        self.entry_usuario.bind("<Return>", self.fazer_login)
        self.entry_senha.bind("<Return>", self.fazer_login)

        btn_entrar = ctk.CTkButton(login_card, text="ACESSAR SISTEMA", command=self.fazer_login, width=300, height=50, font=("Roboto Bold", 15), corner_radius=10)
        btn_entrar.pack(pady=(25, 50), padx=50) 

    def fazer_login(self, event=None):
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()

        if usuario in USUARIOS_PERMITIDOS and USUARIOS_PERMITIDOS[usuario] == senha:
            self.usuario_logado = usuario
            self.tela_principal()
        else:
            messagebox.showerror("Erro de Acesso", "Usuário ou senha incorretos.")

    def fazer_logout(self):
        self.usuario_logado = None
        self.tela_login()

    # ==========================================
    # AUTOCOMPLETAR FLUTUANTE 
    # ==========================================
    def bind_autocomplete(self, combo, get_valores_func, comando_cascata=None):
        janela_principal = combo.winfo_toplevel()
        
        lista_flutuante = tk.Listbox(janela_principal, bg="#ffffff", fg="black", font=("Roboto", 12),
                                     borderwidth=1, relief="solid", selectbackground="#c8e6c9", selectforeground="black", highlightthickness=0)
        
        def atualizar_lista(event):
            if getattr(event, 'keysym', '') in ['Up', 'Down', 'Return', 'Escape', 'Tab']:
                return
            
            texto = combo.get().strip()
            opcoes = get_valores_func()
            
            if not texto:
                lista_flutuante.place_forget()
                return

            filtrado = [op for op in opcoes if texto.lower() in op.lower()]
            
            lista_flutuante.delete(0, tk.END)
            for item in filtrado:
                lista_flutuante.insert(tk.END, item)
                
            if filtrado:
                altura_lista = min(120, len(filtrado) * 25) 
                x_pos = combo.winfo_rootx() - janela_principal.winfo_rootx()
                y_pos = combo.winfo_rooty() - janela_principal.winfo_rooty() + combo.winfo_height()
                
                lista_flutuante.place(x=x_pos, y=y_pos, width=combo.winfo_width(), height=altura_lista)
                lista_flutuante.lift() 
            else:
                lista_flutuante.place_forget()

        def selecionar_item(event):
            if lista_flutuante.curselection():
                item_selecionado = lista_flutuante.get(lista_flutuante.curselection())
                combo.set(item_selecionado) 
                lista_flutuante.place_forget() 
                
                if hasattr(combo, '_entry'):
                    combo._entry.focus_set()
                
                if comando_cascata:
                    comando_cascata(item_selecionado)

        def esconder_lista(event):
            combo.after(200, lambda: lista_flutuante.place_forget())

        entry_widget = combo._entry if hasattr(combo, '_entry') else combo
        entry_widget.bind('<KeyRelease>', atualizar_lista)
        entry_widget.bind('<FocusOut>', esconder_lista)
        lista_flutuante.bind('<<ListboxSelect>>', selecionar_item)

    # --- LÓGICA DE CASCATA ---
    def obter_opcoes(self, alvo, und="", cc="", sub="", gestor="", posto=""):
        resultados = set()
        for linha in self.dados_excel["linhas"]:
            if und and linha["unidade"] != und: continue
            if cc and linha["cc"] != cc: continue
            if sub and linha["sub"] != sub: continue
            if gestor and linha["gestor"] != gestor: continue
            if posto and linha["posto"] != posto: continue
            
            if linha[alvo]:
                resultados.add(linha[alvo])
        return sorted(list(resultados))

    def criar_campo(self, parent, texto_label, row, get_valores_func, comando_cascata=None):
        ctk.CTkLabel(parent, text=texto_label, font=("Roboto", 12)).grid(row=row, column=0, sticky="w", pady=5, padx=10)
        
        valores_iniciais = get_valores_func()
        if not valores_iniciais: valores_iniciais = [""]
            
        combo = ctk.CTkComboBox(parent, values=valores_iniciais, width=340, command=comando_cascata)
        combo.set("") 
        combo.grid(row=row, column=1, pady=5, padx=10)
        
        self.bind_autocomplete(combo, get_valores_func, comando_cascata)
        ToolTip(combo)
        return combo

    # --- ATUALIZADORES DE CASCATA (SAÍDA) ---
    def atualizar_cascata_s_und(self, escolha):
        self.combo_cc_s.set(''); self.combo_sub_s.set(''); self.combo_gestor_s.set(''); self.combo_posto_s.set(''); self.combo_cargo_s.set('')
        self.combo_cc_s.configure(values=self.obter_opcoes("cc", und=escolha))
    def atualizar_cascata_s_cc(self, escolha):
        self.combo_sub_s.set(''); self.combo_gestor_s.set(''); self.combo_posto_s.set(''); self.combo_cargo_s.set('')
        self.combo_sub_s.configure(values=self.obter_opcoes("sub", und=self.combo_und_s.get(), cc=escolha))
    def atualizar_cascata_s_sub(self, escolha):
        self.combo_gestor_s.set(''); self.combo_posto_s.set(''); self.combo_cargo_s.set('')
        self.combo_gestor_s.configure(values=self.obter_opcoes("gestor", und=self.combo_und_s.get(), cc=self.combo_cc_s.get(), sub=escolha))
    def atualizar_cascata_s_gestor(self, escolha):
        self.combo_posto_s.set(''); self.combo_cargo_s.set('')
        self.combo_posto_s.configure(values=self.obter_opcoes("posto", und=self.combo_und_s.get(), cc=self.combo_cc_s.get(), sub=self.combo_sub_s.get(), gestor=escolha))
    def atualizar_cascata_s_posto(self, escolha):
        self.combo_cargo_s.set('')
        self.combo_cargo_s.configure(values=self.obter_opcoes("cargo", und=self.combo_und_s.get(), cc=self.combo_cc_s.get(), sub=self.combo_sub_s.get(), gestor=self.combo_gestor_s.get(), posto=escolha))

    # --- ATUALIZADORES DE CASCATA (ENTRADA) ---
    def atualizar_cascata_e_und(self, escolha):
        self.combo_cc_e.set(''); self.combo_sub_e.set(''); self.combo_gestor_e.set(''); self.combo_posto_e.set(''); self.combo_cargo_e.set('')
        self.combo_cc_e.configure(values=self.obter_opcoes("cc", und=escolha))
    def atualizar_cascata_e_cc(self, escolha):
        self.combo_sub_e.set(''); self.combo_gestor_e.set(''); self.combo_posto_e.set(''); self.combo_cargo_e.set('')
        self.combo_sub_e.configure(values=self.obter_opcoes("sub", und=self.combo_und_e.get(), cc=escolha))
    def atualizar_cascata_e_sub(self, escolha):
        self.combo_gestor_e.set(''); self.combo_posto_e.set(''); self.combo_cargo_e.set('')
        self.combo_gestor_e.configure(values=self.obter_opcoes("gestor", und=self.combo_und_e.get(), cc=self.combo_cc_e.get(), sub=escolha))
    def atualizar_cascata_e_gestor(self, escolha):
        self.combo_posto_e.set(''); self.combo_cargo_e.set('')
        self.combo_posto_e.configure(values=self.obter_opcoes("posto", und=self.combo_und_e.get(), cc=self.combo_cc_e.get(), sub=self.combo_sub_e.get(), gestor=escolha))
    def atualizar_cascata_e_posto(self, escolha):
        self.combo_cargo_e.set('')
        self.combo_cargo_e.configure(values=self.obter_opcoes("cargo", und=self.combo_und_e.get(), cc=self.combo_cc_e.get(), sub=self.combo_sub_e.get(), gestor=self.combo_gestor_e.get(), posto=escolha))

    # ==========================================
    # TELA SECUNDÁRIA: SOLICITAR NOVO POSTO
    # ==========================================
    def abrir_solicitacao_posto(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Solicitação de Novo Posto")
        popup.geometry("550x550") 
        popup.transient(self) 
        popup.grab_set() 
        
        ctk.CTkLabel(popup, text="Avisar RH sobre Posto Faltante", font=("Roboto Bold", 22)).pack(pady=(20, 5))
        ctk.CTkLabel(popup, text="Preencha os dados abaixo. O RH criará a numeração oficial.", font=("Roboto", 13), text_color="gray").pack(pady=(0, 20))
        
        frame_form = ctk.CTkFrame(popup, fg_color="transparent")
        frame_form.pack(fill="both", expand=True, padx=30)
        
        def criar_campo_popup(label_text, get_valores_func):
            ctk.CTkLabel(frame_form, text=label_text, font=("Roboto Bold", 12)).pack(anchor="w", pady=(5,0))
            combo = ctk.CTkComboBox(frame_form, values=get_valores_func(), width=490)
            combo.set("")
            combo.pack(anchor="w", pady=(0, 10))
            self.bind_autocomplete(combo, get_valores_func)
            return combo
            
        combo_und = criar_campo_popup("Unidade:", lambda: self.obter_opcoes("unidade"))
        combo_cc = criar_campo_popup("Centro de Custo:", lambda: self.obter_opcoes("cc"))
        combo_sub = criar_campo_popup("Subprocesso:", lambda: self.obter_opcoes("sub"))
        combo_gestor = criar_campo_popup("Gestor:", lambda: self.obter_opcoes("gestor"))
        
        ctk.CTkLabel(frame_form, text="Qual Cargo deve pertencer a esse posto?:", font=("Roboto Bold", 12), text_color=("#c62828", "#ffcc00")).pack(anchor="w", pady=(5,0))
        combo_cargo = ctk.CTkComboBox(frame_form, values=self.obter_opcoes("cargo"), width=490)
        combo_cargo.set("")
        combo_cargo.pack(anchor="w", pady=(0, 10))
        self.bind_autocomplete(combo_cargo, lambda: self.obter_opcoes("cargo"))
        
        def salvar_solicitacao():
            v_und = combo_und.get().strip()
            v_cc = combo_cc.get().strip()
            v_sub = combo_sub.get().strip()
            v_gestor = combo_gestor.get().strip()
            v_cargo = combo_cargo.get().strip()
            
            if not all([v_und, v_cc, v_sub, v_gestor, v_cargo]):
                messagebox.showwarning("Aviso", "Todos os campos são obrigatórios para a solicitação.", parent=popup)
                return
                
            arquivo_solicitacoes = "solicitacoes_postos.xlsx"
            try:
                if os.path.exists(arquivo_solicitacoes):
                    wb = openpyxl.load_workbook(arquivo_solicitacoes)
                    ws = wb.active
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Solicitações de Postos"
                    ws.append(["Data Solicitação", "Usuário", "Unidade", "Centro de Custo", "Subprocesso", "Gestor", "Cargo"])
                    
                data_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                ws.append([data_atual, self.usuario_logado, v_und, v_cc, v_sub, v_gestor, v_cargo])
                wb.save(arquivo_solicitacoes)
                
                messagebox.showinfo("Sucesso", "Sua solicitação foi enviada com sucesso! O RH foi notificado para criar o posto oficial.", parent=popup)
                popup.destroy() 
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar solicitação. O arquivo pode estar aberto por outra pessoa.\nErro: {e}", parent=popup)

        btn_enviar = ctk.CTkButton(popup, text="ENVIAR SOLICITAÇÃO AO RH", font=("Roboto Bold", 15), height=45, fg_color="#1f538d", hover_color="#153b66", command=salvar_solicitacao)
        btn_enviar.pack(pady=(20, 15))

    # --- TELA PRINCIPAL (REGISTRO) ---
    def tela_principal(self):
        self.limpar_tela()

        frame_top = ctk.CTkFrame(self, fg_color="transparent")
        frame_top.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(frame_top, text="Nova Movimentação", font=("Roboto Bold", 24)).pack(side="left")
        
        btn_sair = ctk.CTkButton(frame_top, text="Sair", width=60, height=35, fg_color="#d32f2f", hover_color="#b71c1c", font=("Roboto Bold", 12), command=self.fazer_logout)
        btn_sair.pack(side="right", padx=(10, 0))
        
        btn_consulta = ctk.CTkButton(frame_top, text="Minhas Movimentações", height=35, fg_color="#1f538d", font=("Roboto Bold", 12), command=self.tela_consulta)
        btn_consulta.pack(side="right", padx=(10, 0))

        user_frame = ctk.CTkFrame(frame_top, fg_color=("#e0e0e0", "#424242"), corner_radius=20)
        user_frame.pack(side="right")
        ctk.CTkLabel(user_frame, text=f"👤 {self.usuario_logado}", font=("Roboto", 12), padx=15, pady=5).pack()

        container_main = ctk.CTkFrame(self, fg_color="transparent")
        container_main.pack(fill="both", expand=True, padx=20, pady=(0, 5))

        # Requisitante
        frame_req = ctk.CTkFrame(container_main)
        frame_req.pack(fill="x", pady=5)
        ctk.CTkLabel(frame_req, text="Quem solicitou a troca?", font=("Roboto Bold", 13)).pack(side="left", padx=(20, 10), pady=10)
        
        valores_req = self.dados_excel["requisitantes"] if self.dados_excel["requisitantes"] else [""]
        self.combo_requisitante = ctk.CTkComboBox(frame_req, values=valores_req, width=500, height=30, font=("Roboto", 13))
        self.combo_requisitante.set("")
        self.combo_requisitante.pack(side="left", padx=10, pady=10)
        
        self.bind_autocomplete(self.combo_requisitante, lambda: self.dados_excel["requisitantes"])
        ToolTip(self.combo_requisitante) 

        frame_split = ctk.CTkFrame(container_main, fg_color="transparent")
        frame_split.pack(fill="both", expand=True)

        # ---- LADO ESQUERDO: SAÍDA ----
        frame_saida = ctk.CTkFrame(frame_split, corner_radius=15, fg_color=("#ffebee", "#3e2723")) 
        frame_saida.pack(side="left", fill="both", expand=True, padx=(0, 10), pady=5)
        
        header_saida = ctk.CTkFrame(frame_saida, fg_color="#c62828", corner_radius=15, height=40)
        header_saida.pack(fill="x")
        ctk.CTkLabel(header_saida, text="VAGA DE SAÍDA (RETIRADA)", font=("Roboto Bold", 13), text_color="white").place(relx=0.5, rely=0.5, anchor="center")
        
        content_saida = ctk.CTkFrame(frame_saida, fg_color="transparent")
        content_saida.pack(fill="both", expand=True, padx=10, pady=10)

        self.combo_und_s = self.criar_campo(content_saida, "Unidade:", 0, lambda: self.obter_opcoes("unidade"), self.atualizar_cascata_s_und)
        self.combo_cc_s = self.criar_campo(content_saida, "Centro de Custo:", 1, lambda: self.obter_opcoes("cc", self.combo_und_s.get()), self.atualizar_cascata_s_cc)
        self.combo_sub_s = self.criar_campo(content_saida, "Subprocesso:", 2, lambda: self.obter_opcoes("sub", self.combo_und_s.get(), self.combo_cc_s.get()), self.atualizar_cascata_s_sub)
        self.combo_gestor_s = self.criar_campo(content_saida, "Gestor:", 3, lambda: self.obter_opcoes("gestor", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get()), self.atualizar_cascata_s_gestor)
        self.combo_posto_s = self.criar_campo(content_saida, "Posto:", 4, lambda: self.obter_opcoes("posto", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get(), self.combo_gestor_s.get()), self.atualizar_cascata_s_posto)
        self.combo_cargo_s = self.criar_campo(content_saida, "Cargo:", 5, lambda: self.obter_opcoes("cargo", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get(), self.combo_gestor_s.get(), self.combo_posto_s.get()))

        ctk.CTkLabel(content_saida, text="Quantidade:", font=("Roboto Bold", 12)).grid(row=6, column=0, sticky="w", pady=(15,5), padx=10)
        frame_qtd_s = ctk.CTkFrame(content_saida, fg_color=("#ffcdd2", "#5c3632"), corner_radius=10)
        frame_qtd_s.grid(row=6, column=1, sticky="w", padx=10, pady=(15,5))
        ctk.CTkButton(frame_qtd_s, text="-", width=35, height=30, fg_color="transparent", text_color=("black", "white"), hover_color=("#ef9a9a", "#4e342e"), font=("Roboto Bold", 18), command=lambda: self.alterar_qtd(self.qtd_saida, -1)).pack(side="left")
        ctk.CTkLabel(frame_qtd_s, textvariable=self.qtd_saida, font=("Roboto Bold", 16), width=35).pack(side="left")
        ctk.CTkButton(frame_qtd_s, text="+", width=35, height=30, fg_color="transparent", text_color=("black", "white"), hover_color=("#ef9a9a", "#4e342e"), font=("Roboto Bold", 18), command=lambda: self.alterar_qtd(self.qtd_saida, 1)).pack(side="left")

        # ---- LADO DIREITO: ENTRADA ----
        frame_entrada = ctk.CTkFrame(frame_split, corner_radius=15, fg_color=("#e8f5e9", "#1b3320"))
        frame_entrada.pack(side="right", fill="both", expand=True, padx=(10, 0), pady=5)
        
        header_entrada = ctk.CTkFrame(frame_entrada, fg_color="#2e7d32", corner_radius=15, height=40)
        header_entrada.pack(fill="x")
        ctk.CTkLabel(header_entrada, text="VAGA DE ENTRADA (NOVA)", font=("Roboto Bold", 13), text_color="white").place(relx=0.5, rely=0.5, anchor="center")
        
        content_entrada = ctk.CTkFrame(frame_entrada, fg_color="transparent")
        content_entrada.pack(fill="both", expand=True, padx=10, pady=10)

        self.combo_und_e = self.criar_campo(content_entrada, "Unidade:", 0, lambda: self.obter_opcoes("unidade"), self.atualizar_cascata_e_und)
        self.combo_cc_e = self.criar_campo(content_entrada, "Centro de Custo:", 1, lambda: self.obter_opcoes("cc", self.combo_und_e.get()), self.atualizar_cascata_e_cc)
        self.combo_sub_e = self.criar_campo(content_entrada, "Subprocesso:", 2, lambda: self.obter_opcoes("sub", self.combo_und_e.get(), self.combo_cc_e.get()), self.atualizar_cascata_e_sub)
        self.combo_gestor_e = self.criar_campo(content_entrada, "Gestor:", 3, lambda: self.obter_opcoes("gestor", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get()), self.atualizar_cascata_e_gestor)
        self.combo_posto_e = self.criar_campo(content_entrada, "Posto:", 4, lambda: self.obter_opcoes("posto", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get(), self.combo_gestor_e.get()), self.atualizar_cascata_e_posto)
        self.combo_cargo_e = self.criar_campo(content_entrada, "Cargo:", 5, lambda: self.obter_opcoes("cargo", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get(), self.combo_gestor_e.get(), self.combo_posto_e.get()))

        ctk.CTkLabel(content_entrada, text="Quantidade:", font=("Roboto Bold", 12)).grid(row=6, column=0, sticky="w", pady=(15,5), padx=10)
        frame_qtd_e = ctk.CTkFrame(content_entrada, fg_color=("#c8e6c9", "#2a472e"), corner_radius=10)
        frame_qtd_e.grid(row=6, column=1, sticky="w", padx=10, pady=(15,5))
        ctk.CTkButton(frame_qtd_e, text="-", width=35, height=30, fg_color="transparent", text_color=("black", "white"), hover_color=("#a5d6a7", "#385f3c"), font=("Roboto Bold", 18), command=lambda: self.alterar_qtd(self.qtd_entrada, -1)).pack(side="left")
        ctk.CTkLabel(frame_qtd_e, textvariable=self.qtd_entrada, font=("Roboto Bold", 16), width=35).pack(side="left")
        ctk.CTkButton(frame_qtd_e, text="+", width=35, height=30, fg_color="transparent", text_color=("black", "white"), hover_color=("#a5d6a7", "#385f3c"), font=("Roboto Bold", 18), command=lambda: self.alterar_qtd(self.qtd_entrada, 1)).pack(side="left")

        # LINK PARA SOLICITAR POSTO NOVO NO LUGAR CERTO E ELEGANTE
        btn_falta_posto = ctk.CTkButton(content_entrada, text="Não encontrou o posto? Clique aqui para solicitar cadastro", 
                                        fg_color="transparent", text_color=("#c62828", "#ffcc00"), hover_color=("#e8f5e9", "#1b3320"), 
                                        font=("Roboto", 12, "underline"), command=self.abrir_solicitacao_posto)
        btn_falta_posto.grid(row=7, column=0, columnspan=2, pady=(20, 0))

        # ---- BOTÃO SALVAR ----
        btn_salvar = ctk.CTkButton(container_main, text="CONFIRMAR MOVIMENTAÇÃO", font=("Roboto Bold", 15), height=45, width=350, corner_radius=10, command=self.salvar_dados)
        btn_salvar.pack(pady=(15, 10))

    def alterar_qtd(self, variavel, valor):
        nova_qtd = variavel.get() + valor
        if nova_qtd >= 1:
            variavel.set(nova_qtd)

    def salvar_dados(self):
        combos_validar = [
            (self.combo_requisitante, "Requisitante", lambda: self.dados_excel["requisitantes"]),
            (self.combo_und_s, "Unidade (Saída)", lambda: self.obter_opcoes("unidade")), 
            (self.combo_cc_s, "Centro de Custo (Saída)", lambda: self.obter_opcoes("cc", self.combo_und_s.get())), 
            (self.combo_sub_s, "Subprocesso (Saída)", lambda: self.obter_opcoes("sub", self.combo_und_s.get(), self.combo_cc_s.get())), 
            (self.combo_gestor_s, "Gestor (Saída)", lambda: self.obter_opcoes("gestor", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get())), 
            (self.combo_posto_s, "Posto (Saída)", lambda: self.obter_opcoes("posto", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get(), self.combo_gestor_s.get())), 
            (self.combo_cargo_s, "Cargo (Saída)", lambda: self.obter_opcoes("cargo", self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get(), self.combo_gestor_s.get(), self.combo_posto_s.get())),
            (self.combo_und_e, "Unidade (Entrada)", lambda: self.obter_opcoes("unidade")), 
            (self.combo_cc_e, "Centro de Custo (Entrada)", lambda: self.obter_opcoes("cc", self.combo_und_e.get())), 
            (self.combo_sub_e, "Subprocesso (Entrada)", lambda: self.obter_opcoes("sub", self.combo_und_e.get(), self.combo_cc_e.get())), 
            (self.combo_gestor_e, "Gestor (Entrada)", lambda: self.obter_opcoes("gestor", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get())), 
            (self.combo_posto_e, "Posto (Entrada)", lambda: self.obter_opcoes("posto", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get(), self.combo_gestor_e.get())), 
            (self.combo_cargo_e, "Cargo (Entrada)", lambda: self.obter_opcoes("cargo", self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get(), self.combo_gestor_e.get(), self.combo_posto_e.get()))
        ]

        # VALIDAÇÃO
        for combo, nome_campo, func_valores in combos_validar:
            valor = combo.get().strip()
            if valor == "":
                messagebox.showwarning("Aviso", f"O campo '{nome_campo}' é obrigatório.")
                return
            
            opcoes_validas = func_valores()
            if valor not in opcoes_validas:
                messagebox.showwarning("Aviso", f"Atenção no campo '{nome_campo}'!\nVocê digitou '{valor}', mas precisa clicar na lista suspensa para selecionar o nome completo.")
                return

        conn = conectar_banco()
        cursor = conn.cursor()
        data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        cursor.execute('''
            INSERT INTO movimentacoes (
                usuario_sistema, data_registro, requisitante,
                unidade_saida, cc_saida, subprocesso_saida, gestor_saida, posto_saida, cargo_saida, qtd_saida,
                unidade_entrada, cc_entrada, subprocesso_entrada, gestor_entrada, posto_entrada, cargo_entrada, qtd_entrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            self.usuario_logado, data_atual, self.combo_requisitante.get(),
            self.combo_und_s.get(), self.combo_cc_s.get(), self.combo_sub_s.get(), self.combo_gestor_s.get(), self.combo_posto_s.get(), self.combo_cargo_s.get(), self.qtd_saida.get(),
            self.combo_und_e.get(), self.combo_cc_e.get(), self.combo_sub_e.get(), self.combo_gestor_e.get(), self.combo_posto_e.get(), self.combo_cargo_e.get(), self.qtd_entrada.get()
        ))
        
        conn.commit()

        # ESPELHO EM EXCEL
        try:
            cursor.execute("SELECT * FROM movimentacoes")
            dados_completos = cursor.fetchall()
            colunas = [descricao[0] for descricao in cursor.description]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Base_PowerBI"

            ws.append(colunas)
            for linha_banco in dados_completos:
                ws.append(linha_banco)

            wb.save('base_powerbi.xlsx')
        except Exception as e:
            print(f"Erro ao gerar espelho Excel: {e}")

        conn.close()

        messagebox.showinfo("Sucesso", "Movimentação registrada com sucesso!")
        
        self.combo_requisitante.set('')
        for combo, _, _ in combos_validar[1:]:
            combo.set('')
        self.qtd_saida.set(1)
        self.qtd_entrada.set(1)

    # --- TELA DE CONSULTA DE HISTÓRICO (Com visual Light adaptado) ---
    def tela_consulta(self):
        self.limpar_tela()

        frame_top = ctk.CTkFrame(self, fg_color="transparent")
        frame_top.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(frame_top, text="Histórico de Movimentações", font=("Roboto Bold", 26)).pack(side="left")
        
        btn_voltar = ctk.CTkButton(frame_top, text="Voltar para Registro", width=150, height=35, fg_color="#1f538d", font=("Roboto Bold", 13), command=self.tela_principal)
        btn_voltar.pack(side="right")

        conn = conectar_banco()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, data_registro, requisitante, qtd_saida, cargo_saida, qtd_entrada, cargo_entrada 
            FROM movimentacoes WHERE usuario_sistema = ? ORDER BY id DESC
        ''', (self.usuario_logado,))
        registros = cursor.fetchall()
        conn.close()

        total_registros = len(registros)
        ultima_data = "-"
        if total_registros > 0:
            data_str = registros[0][1]
            try:
                dt = datetime.strptime(data_str, "%Y-%m-%d %H:%M:%S")
                ultima_data = dt.strftime("%d/%m/%Y às %H:%M")
            except:
                ultima_data = data_str

        cards_frame = ctk.CTkFrame(self, fg_color="transparent")
        cards_frame.pack(fill="x", padx=20, pady=(0, 20))

        card_total = ctk.CTkFrame(cards_frame, fg_color=("#e0e0e0", "#333333"), corner_radius=15)
        card_total.pack(side="left", fill="both", expand=True, padx=(0, 10))
        ctk.CTkLabel(card_total, text="TOTAL REGISTRADO", font=("Roboto Bold", 12), text_color="#707070").pack(pady=(15, 0))
        ctk.CTkLabel(card_total, text=str(total_registros), font=("Roboto Bold", 36), text_color="#2e7d32").pack(pady=(0, 15))

        card_data = ctk.CTkFrame(cards_frame, fg_color=("#e0e0e0", "#333333"), corner_radius=15)
        card_data.pack(side="left", fill="both", expand=True, padx=(10, 0))
        ctk.CTkLabel(card_data, text="ÚLTIMA MOVIMENTAÇÃO", font=("Roboto Bold", 12), text_color="#707070").pack(pady=(15, 0))
        ctk.CTkLabel(card_data, text=ultima_data, font=("Roboto Bold", 24)).pack(pady=(10, 15))

        # CORES DA TABELA ATUALIZADAS PARA MODO CLARO
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#ffffff", foreground="black", fieldbackground="#ffffff", borderwidth=0, rowheight=35, font=("Roboto", 11))
        style.configure("Treeview.Heading", background="#2e7d32", foreground="white", font=("Roboto Bold", 12), relief="flat", padding=5)
        style.map('Treeview', background=[('selected', '#c8e6c9')], foreground=[('selected', 'black')]) 

        frame_tabela = ctk.CTkFrame(self, corner_radius=15)
        frame_tabela.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        colunas = ("ID", "Data", "Requisitante", "Qtd Saída", "Cargo Saída", "Qtd Entrada", "Cargo Entrada")
        tabela = ttk.Treeview(frame_tabela, columns=colunas, show="headings", style="Treeview")

        larguras = [40, 140, 160, 80, 180, 80, 180]
        for col, larg in zip(colunas, larguras):
            tabela.heading(col, text=col.upper())
            tabela.column(col, width=larg, anchor="center")

        tabela.pack(side="left", fill="both", expand=True, padx=2, pady=2) 

        scrollbar = ttk.Scrollbar(frame_tabela, orient="vertical", command=tabela.yview)
        tabela.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y", pady=2)

        # LINHAS ALTERNADAS CLARAS
        tabela.tag_configure("par", background="#f5f5f5")
        tabela.tag_configure("impar", background="#ffffff")

        for i, linha in enumerate(registros):
            linha_lista = list(linha)
            try:
                dt_obj = datetime.strptime(linha_lista[1], "%Y-%m-%d %H:%M:%S")
                linha_lista[1] = dt_obj.strftime("%d/%m/%Y %H:%M")
            except:
                pass
            
            tag = "par" if i % 2 == 0 else "impar"
            tabela.insert("", "end", values=linha_lista, tags=(tag,))

# ==========================================
# 5. INICIALIZAÇÃO
# ==========================================
if __name__ == "__main__":
    conectar_banco()
    app = AppHeadcount()
    app.mainloop()